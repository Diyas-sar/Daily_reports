import openpyxl
from datetime import datetime, timedelta
import os
from urllib.request import urlretrieve
import time
import win32com.client as win32
import xlrd
import webbrowser
from win32com import client
# Record the start time
start_time = time.time()


# Get yesterday's date in the desired format
date = (datetime.now() - timedelta(1)).strftime("%d.%m.")

# File paths
final_path = "source file path"
excel_file_url = 'source file path'
excel_file_path = "C:\\Users\\diasasar\\Downloads\\DailyReport2CDProduct.Dashb..xlsx"

excel_67_url = 'source file path'
excel_67_file_path = "C:\\Users\\diasasar\\Downloads\\Form67.xlsx"

stock_value_url = 'source file path'
stock_value_path ='C:\\Users\\diasasar\\Downloads\\DailyReport.xls'

# Load the final workbook
final_workbook = openpyxl.load_workbook(final_path)


# Download the Excel file
try:
    # Remove the existing file if it exists
    if os.path.exists(excel_67_file_path):
        os.remove(excel_67_file_path)

    # Download the new file
    urlretrieve(excel_67_url, excel_67_file_path)
    print("Форма 67 загружена")
except Exception as e:
    print("Ошибка с загрузкой 67 формы, проверь интернет:", str(e))

try:
    # Remove the existing file if it exists
    if os.path.exists(excel_file_path):
        os.remove(excel_file_path)

    # Download the new file
    urlretrieve(excel_file_url, excel_file_path)
    print("Ежедневный отчет загружен")
except Exception as e:
    print("Ошибка загрузки ежедневного отчета, проверь подключение:", str(e))

try:
    # Remove the existing file if it exists
    if os.path.exists(stock_value_path):
        os.remove(stock_value_path)

    # Download the new file
    urlretrieve(stock_value_url,stock_value_path)
    print("Остатки загружены")
except Exception as e:
    print("Ошибка с загрузкой Остатков, проверь интернет:", str(e))

# Open the Excel file
webbrowser.open(excel_file_path)
print("Excel file opened successfully!")

# Select the sheet for the given date
final_sheet = final_workbook[date]

# Unhide the sheet if hidden
if final_sheet.sheet_state == 'hidden':
    final_sheet.sheet_state = 'visible'

# Save the changes to the final workbook
final_workbook.save(final_path)

# Load the source and final workbooks
source_workbook = openpyxl.load_workbook(excel_file_path)
final_workbook = openpyxl.load_workbook(final_path)
form_67 = openpyxl.load_workbook(excel_67_file_path)

# Select the source and final sheets
source_sheet = source_workbook['Daily report2 CD Product.Dashb.']
final_sheet = final_workbook[date]
sheet_67 = form_67['Оперативный план']


# Copy the data from source to final

final_sheet['I33'].value=sheet_67['F8'].value/100
final_sheet['M33'].value=sheet_67['N8'].value/100


for row in range(13, 30):
    for col in range(8, 10):
        source_cell = source_sheet.cell(row=row, column=col)
        final_cell = final_sheet.cell(row=row, column=col)
        final_cell.value = source_cell.value
        final_cell.number_format = source_cell.number_format


for row in range(13, 30):
    for col in range(12, 14):
        source_cell = source_sheet.cell(row=row, column=col)
        final_cell = final_sheet.cell(row=row, column=col)
        final_cell.value = source_cell.value
        final_cell.number_format = source_cell.number_format


for row in range(39, 74):
    for col in range(8, 14):
        source_cell = source_sheet.cell(row=row, column=col)
        final_cell = final_sheet.cell(row=row, column=col)
        final_cell.value = source_cell.value
        final_cell.number_format = source_cell.number_format

for row in range(79, 86):
    for col in range(8, 14):
        source_cell = source_sheet.cell(row=row, column=col)
        final_cell = final_sheet.cell(row=row, column=col)
        final_cell.value = source_cell.value
        final_cell.number_format = source_cell.number_format





#Скачать файл- источник из почты


path = r"C:\\Users\\diasasar\\Desktop\\Daily_reports"
outlook = win32.Dispatch("Outlook.Application").GetNamespace("MAPI")
inbox = outlook.GetDefaultFolder(6)
messages = inbox.Items
message = messages.GetLast()

def save_attachments(subject_prefix):
    messages.Sort("[ReceivedTime]", True)  # sort by received date: newest to oldest
    for message in messages:
        if message.Subject.startswith(subject_prefix):
            print("saving attachments for:", message.Subject)
            for attachment in message.Attachments:
                file_path = os.path.join(path, str(attachment.FileName))
                if os.path.exists(file_path):
                    os.remove(file_path)  # Delete the file if it exists
                attachment.SaveAsFile(file_path)
                print(attachment.FileName, "saved.")
            return  # exit after first matched message


save_attachments('FW: Отправка: Справка о зольности')



# Path to the .xls file
xls_file_path = "C:\\Users\\diasasar\\Desktop\\Daily_reports\\Справка о зольности  добытых и отгруженных углей  за 29.06.2023г.xls"

# Load the workbook using xlrd
workbook = xlrd.open_workbook(xls_file_path)
#трансформация файла с остатками
workbook_stock = xlrd.open_workbook(stock_value_path)

# Select the sheet you want to read
sheet = workbook.sheet_by_name("Зольность доб. и отгр.углей")
sheet_stock = workbook_stock.sheet_by_index(0)

# Access the data from the sheet
value_z7 = sheet.cell_value(6, 25)  # Assuming cell Z7 contains the desired value
value_z8 = sheet.cell_value(7, 25)  # Assuming cell Z8 contains the desired value
value_z9 = sheet.cell_value(8, 25)  # Assuming cell Z9 contains the desired value
value_z10 = sheet.cell_value(6, 14)  # Assuming cell Z7 contains the desired value
value_z11 = sheet.cell_value(7, 14)  # Assuming cell Z8 contains the desired value
value_z12 = sheet.cell_value(8, 14)  # Assuming cell Z9 contains the desired value

final_sheet['I34'].value=value_z7/100
final_sheet['I35'].value=value_z8/100
final_sheet['I36'].value=value_z9/100
final_sheet['M34'].value=value_z7/100
final_sheet['M35'].value=value_z8/100
final_sheet['M36'].value=value_z9/100

# Access data from the sheet_stock
stock= sheet_stock.cell_value(14, 4)
stock1= sheet_stock.cell_value(15, 4)
stock2= sheet_stock.cell_value(16, 4)
stock3= sheet_stock.cell_value(17, 4)
stock4= sheet_stock.cell_value(18, 4)
stock5= sheet_stock.cell_value(19, 4)
stock6= sheet_stock.cell_value(20, 4)
stock7= sheet_stock.cell_value(21, 4)

stock8= sheet_stock.cell_value(25, 4)
stock9= sheet_stock.cell_value(26, 4)

final_sheet['e23'].value=stock /1000
final_sheet['e24'].value=stock1/1000
final_sheet['e25'].value=stock2/1000
final_sheet['e26'].value=stock3/1000
final_sheet['e27'].value=stock4/1000
final_sheet['e28'].value=stock5 /1000
final_sheet['e29'].value=stock6 /1000
final_sheet['e30'].value=stock7 /1000
final_sheet['e44'].value=stock8 /1000
final_sheet['e48'].value=stock9 /1000

date1 = (datetime.now() - timedelta(1)).strftime("%d_%m_%Y")
# Save the changes to the final workbook
final_workbook.save(f"C:\\Users\\diasasar\\Desktop\\Daily_reports\\Daily report_CD_{date1}.xlsx")


# Record the end time
end_time = time.time()

# Calculate the elapsed time
elapsed_time = end_time - start_time

print("Execution time:", int(elapsed_time), "seconds")

time.sleep(10)
#конвертация файла в pdf

def send_email(subject, body, recipients, attachments=[]):
    try:
        outlook = win32.Dispatch("Outlook.Application")
        mail = outlook.CreateItem(0)  # 0 represents a new mail item

        mail.To = ";".join(recipients) if isinstance(recipients, list) else recipients
        mail.Subject = subject
        mail.Body = body

        for attachment in attachments:
            mail.Attachments.Add(Source=attachment)

        #send the email immediately without displaying it.
            mail.Send()

        # Uncomment the lines below if you want to display the email before sending.
        # mail.Display()
        # print("Email displayed. Please review and send manually.")

    except Exception as e:
        print(f"Error: {e}")
#convert excel to pdf file


time.sleep(20)



if __name__ == "__main__":
    recipients_list = ["dias.sarsenbay@arcelormittal.com"]
    email_subject = f"Daily report_CD_{date1}"
    email_body = "This is a test email with attachments sent using Python and Outlook."

    # Replace the file paths with the actual paths of your files.
    attachments_list = [f"C:\\Users\\diasasar\\Desktop\\Daily_reports\\Daily report_CD_{date1}.xlsx"]

    send_email(email_subject, email_body, recipients_list, attachments_list)




